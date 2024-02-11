import csv
import json
import re
import sys
from openpyxl import Workbook, load_workbook
import argparse


class Lesson:
    def __init__(self, code, title, teacher, teaching_slots) -> None:
        self.code = code
        self.title = title
        self.teacher = teacher

        self.teaching_slots = teaching_slots


class TimeSlot():
    def __init__(self, start, end, classroom, comment) -> None:
        self.start = start
        self.end = end
        self.classroom = classroom

        if comment is not None:
            self.comment = comment


def parse_timeslot(slot):
    timeslot_regex_w_comment = re.compile(
        r"(\d+)-(\d+) *([^\(\)]*?) *(\([^\(\)]*\))")
    timeslot_regex_wo_comment = re.compile(r"(\d+)-(\d+) *([^\(\)]*)")

    matched = timeslot_regex_w_comment.match(slot)
    if matched:
        comment = matched[4]
    else:
        matched = timeslot_regex_wo_comment.match(slot)
        comment = None

    assert matched

    start = int(matched[1])
    if not 10 <= start <= 12:
        start += 12

    end = int(matched[2])
    if not 10 <= end <= 12:
        end += 12

    classroom = matched[3]

    return TimeSlot(start, end, classroom, comment)


def convert_lessons_csv(reader: csv.DictReader):
    day_greek_to_eng = {'ΔΕ': "monday",
                        'ΤΡ': "tuesday",
                        'ΤΕ': "wednesday",
                        'ΠΕ': "thursday",
                        'ΠΑ': "friday"}
    a: list[Lesson] = []
    for lesson in reader:
        teaching_slot = {}
        for day_name_gr, day_name_eng in day_greek_to_eng.items():
            if lesson[day_name_gr].strip():
                teaching_slot[day_name_eng] = parse_timeslot(
                    lesson[day_name_gr].strip())

        a.append(Lesson(lesson['Κωδικός'], lesson['Τίτλος'],
                        lesson['ΔΙΔΑΣΚΩΝ_ΕΠΙΘΕΤΟ'], teaching_slot))

    return a


def convert_lessons_xlsx(wb: Workbook):
    worksheet = wb.active
    column_name = {}
    a: list[Lesson] = []

    day_greek_to_eng = {'ΔΕ': "monday",
                        'ΤΡ': "tuesday",
                        'ΤΕ': "wednesday",
                        'ΠΕ': "thursday",
                        'ΠΑ': "friday"}

    for i, column in enumerate(worksheet.iter_cols()):
        column_name[column[0].value] = i

    for row in worksheet.iter_rows(min_row=2):
        teaching_slot = {}
        for day_name_gr, day_name_eng in day_greek_to_eng.items():
            if row[column_name[day_name_gr]].value.strip():
                teaching_slot[day_name_eng] = parse_timeslot(
                    row[column_name[day_name_gr]].value.strip())
        a.append(Lesson(row[column_name['Κωδικός']].value, row[column_name['Τίτλος']].value,
                        row[column_name['ΔΙΔΑΣΚΩΝ_ΕΠΙΘΕΤΟ']].value, teaching_slot))

    return a

class TimetableEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, Lesson):
            return o.__dict__
        elif isinstance(o, TimeSlot):
            return o.__dict__
        return json.JSONEncoder.default(self, o)


def export_to_json(lesson_info: list[Lesson], version: str):
    return json.dumps({"version": version, "schedule": lesson_info}, ensure_ascii=False, cls=TimetableEncoder, indent=4)


def parse_args(args: list[str]):
    parser = argparse.ArgumentParser()
    parser.add_argument("-v", "--version",
                        help="The version of the program", required=True)
    parser.add_argument(
        "-s", "--schedule", help="The path to the schedule file", required=True)
    parser.add_argument(
        "-j", "--json-file", help="The path to the output file", required=True, type=argparse.FileType('w'))
    parser.add_argument(
        "-t", "--type", help="The schedule file type. Only .csv, .xlsx currently supported", choices=['csv', 'xlsx'], required=True)

    return parser.parse_args(args)


def main(args: list[str]):
    parsed_arguments = parse_args(args)

    match parsed_arguments.type:
        case "csv":
            with open(parsed_arguments.schedule,  mode='r', encoding='utf-8-sig') as csvfile:
                reader = csv.DictReader(csvfile)

                # Convert from CSV format to lessons
                schedule = convert_lessons_csv(reader)
        case "xlsx":
            wb = load_workbook(parsed_arguments.schedule)

            schedule = convert_lessons_xlsx(wb)
        case _:
            assert False

    json = export_to_json(schedule, parsed_arguments.version)

    parsed_arguments.json_file.write(json)


if __name__ == "__main__":
    main(sys.argv[1:])
